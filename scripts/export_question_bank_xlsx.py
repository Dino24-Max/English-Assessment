#!/usr/bin/env python3
"""
Export all questions from the database to an Excel workbook matching
`question_bank_import_template.xlsx` column layout.

Run from project root: python scripts/export_question_bank_xlsx.py
Output: src/main/python/data/question_bank_full_export.xlsx
"""

from __future__ import annotations

import asyncio
import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

project_root = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(project_root / "src" / "main" / "python"))

DATA_DIR = project_root / "src" / "main" / "python" / "data"
OUT_PATH = DATA_DIR / "question_bank_full_export.xlsx"

# Same order as scripts/generate_question_import_template.py
HEADERS = [
    "external_id",
    "module_type",
    "division",
    "department",
    "question_type",
    "question_text",
    "options",
    "correct_answer",
    "points",
    "difficulty_level",
    "cefr_level",
    "is_safety_related",
    "scenario_id",
    "scenario_description",
    "audio_file_path",
    "audio_text",
    "audio_context",
    "question_metadata",
    "note",
]


def _format_options(opt: Any) -> str:
    if opt is None:
        return ""
    if isinstance(opt, str):
        return opt
    if isinstance(opt, list):
        if opt and all(isinstance(x, str) for x in opt):
            return "|".join(opt)
        return json.dumps(opt, ensure_ascii=False)
    if isinstance(opt, dict):
        return json.dumps(opt, ensure_ascii=False)
    return str(opt)


def _enum_value(v: Any) -> str:
    if v is None:
        return ""
    return getattr(v, "value", str(v))


def _row_from_question(q: Any) -> list[Any]:
    meta = q.question_metadata
    if not isinstance(meta, dict):
        meta = {}
    else:
        meta = dict(meta)

    audio_text = meta.get("audio_text") or ""
    audio_context = meta.get("audio_context") or ""
    meta_json = json.dumps(meta, ensure_ascii=False) if meta else ""

    return [
        str(q.id),
        _enum_value(q.module_type),
        _enum_value(q.division),
        q.department or "",
        _enum_value(q.question_type),
        q.question_text or "",
        _format_options(q.options),
        q.correct_answer or "",
        q.points,
        q.difficulty_level,
        q.cefr_level or "",
        "TRUE" if q.is_safety_related else "FALSE",
        q.scenario_id if q.scenario_id is not None else "",
        q.scenario_description or "",
        q.audio_file_path or "",
        audio_text,
        audio_context,
        meta_json,
        "",
    ]


async def export_all() -> int:
    from sqlalchemy import select

    from core.database import async_session_maker
    from models.assessment import Question

    async with async_session_maker() as session:
        result = await session.execute(select(Question).order_by(Question.id))
        questions = result.scalars().all()

    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "说明"
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws_info["A1"] = (
        f"全量题库导出（与 question_bank_import_template 列一致）\n\n"
        f"生成时间：{ts}\n"
        f"题目数量：{len(questions)}\n\n"
        f"• external_id 列为数据库主键 id，便于对照。\n"
        f"• 含 speaking 模块；若只要非口语题可自行筛选 module_type。\n"
    )
    ws_info["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_info.column_dimensions["A"].width = 88

    ws = wb.create_sheet("questions", 1)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    for col, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for r, q in enumerate(questions, start=2):
        for c, val in enumerate(_row_from_question(q), start=1):
            ws.cell(row=r, column=c, value=val)

    ws.freeze_panes = "A2"
    widths = {
        1: 12,
        2: 14,
        3: 10,
        4: 18,
        5: 18,
        6: 50,
        7: 40,
        8: 18,
        9: 8,
        10: 12,
        11: 8,
        12: 16,
        13: 14,
        14: 28,
        15: 28,
        16: 32,
        17: 20,
        18: 36,
        19: 12,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    return len(questions)


def main() -> None:
    n = asyncio.run(export_all())
    print(f"Exported {n} questions to {OUT_PATH}")


if __name__ == "__main__":
    main()
