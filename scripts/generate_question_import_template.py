#!/usr/bin/env python3
"""Generate Excel template for adding question-bank rows (all modules except speaking)."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "src" / "main" / "python" / "data" / "question_bank_import_template.xlsx"

HEADERS = [
    "external_id",
    "module_type",
    "division",
    "department",
    "question_type",
    "question_text",
    "options",
    "correct_answer",
    "points",
    "difficulty_level",
    "cefr_level",
    "is_safety_related",
    "scenario_id",
    "scenario_description",
    "audio_file_path",
    "audio_text",
    "audio_context",
    "question_metadata",
    "note",
]

MODULES_NO_SPEAKING = "listening,time_numbers,grammar,vocabulary,reading"
DIVISIONS = "hotel,marine,casino"
QUESTION_TYPES_NO_SPEAKING = "multiple_choice,fill_blank,category_match,title_selection"

INSTRUCTIONS_ZH = """题库导入模板（不含口语 speaking）

一、用途
  在「questions」表填写新题；保存为 .xlsx 后由导入脚本写入数据库（与 JSON 字段一致）。

二、本模板范围
  • 模块 module_type：仅 listening / time_numbers / grammar / vocabulary / reading（不要填 speaking）
  • 题型 question_type：multiple_choice / fill_blank / category_match / title_selection（不要填 speaking_response）

三、列说明
  • external_id：可选，作者侧唯一 ID（如 author_001），便于追踪
  • module_type / division：必填；division 为 hotel / marine / casino
  • department：可选，如 housekeeping / front_office（与系统部门枚举一致时更佳）
  • question_text：题干
  • options：选择题用 | 分隔选项，如 A|B|C|D；填空/阅读等可留空
  • correct_answer：标准答案（选择题填选项字母或完整选项文本，与现有题库一致即可）
  • points：默认 4；difficulty_level：1–3；cefr_level：如 A2/B1
  • is_safety_related：TRUE / FALSE
  • scenario_id / scenario_description：情景题用
  • audio_file_path / audio_text / audio_context：听力用；路径相对 static/audio 等按现有规范
  • question_metadata：可选 JSON 字符串，如阅读 passage、词汇 terms/definitions 等（与 JSON 题库结构一致）
  • note：仅备注，不导入数据库

四、示例行可删除；从第 2 行起追加你的题目。
"""


def main() -> None:
    wb = Workbook()

    # --- Sheet: 说明 ---
    ws_info = wb.active
    ws_info.title = "说明"
    ws_info["A1"] = INSTRUCTIONS_ZH
    ws_info["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_info.column_dimensions["A"].width = 100
    ws_info.row_dimensions[1].height = 420

    # --- Sheet: questions ---
    ws = wb.create_sheet("questions", 1)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    for col, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    examples = [
        [
            "example_listening_001",
            "listening",
            "hotel",
            "front_office",
            "multiple_choice",
            "Guest: I need a wake-up call at 7. Staff: ______",
            "A) Of course, sir.|B) No problem.|C) I don't know.|D) Maybe later.",
            "B",
            4,
            2,
            "A2",
            "FALSE",
            "",
            "",
            "static/audio/wake_up_call.mp3",
            "Guest requests 7am wake-up call.",
            "Front desk dialogue",
            '{"audio_text":"...","audio_context":"..."}',
            "示例：听力选择题，可删",
        ],
        [
            "example_grammar_001",
            "grammar",
            "marine",
            "deck",
            "multiple_choice",
            "The crew ______ the safety drill yesterday.",
            "A) completes|B) completed|C) completing|D) complete",
            "B",
            4,
            2,
            "B1",
            "TRUE",
            "",
            "",
            "",
            "",
            "",
            "",
            "示例：语法，可删",
        ],
        [
            "example_vocab_001",
            "vocabulary",
            "casino",
            "",
            "category_match",
            "Match each term with its definition.",
            '{"terms":["rake","pit"],"definitions":["House commission","Table area"]}',
            '{"rake":"House commission","pit":"Table area"}',
            4,
            2,
            "B1",
            "FALSE",
            "",
            "",
            "",
            "",
            "",
            '{"terms":["rake","pit"],"definitions":["..."],"correct_matches":{}}',
            "示例：词汇配对，JSON 按题库格式",
        ],
    ]

    for r, row in enumerate(examples, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)

    ws.freeze_panes = "A2"
    widths = {
        1: 22,
        2: 14,
        3: 10,
        4: 16,
        5: 18,
        6: 50,
        7: 40,
        8: 18,
        9: 8,
        10: 12,
        11: 8,
        12: 16,
        13: 14,
        14: 28,
        15: 28,
        16: 32,
        17: 20,
        18: 36,
        19: 24,
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    last_row = 5000
    dv_mod = DataValidation(
        type="list",
        formula1=f'"{MODULES_NO_SPEAKING}"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="无效模块",
        error="请从下拉中选择：listening, time_numbers, grammar, vocabulary, reading（不要 speaking）",
    )
    dv_div = DataValidation(
        type="list",
        formula1=f'"{DIVISIONS}"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="无效事业部",
        error="请选择 hotel / marine / casino",
    )
    dv_qt = DataValidation(
        type="list",
        formula1=f'"{QUESTION_TYPES_NO_SPEAKING}"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="无效题型",
        error="请选择 multiple_choice / fill_blank / category_match / title_selection",
    )
    for dv in (dv_mod, dv_div, dv_qt):
        ws.add_data_validation(dv)

    dv_mod.add(f"B2:B{last_row}")
    dv_div.add(f"C2:C{last_row}")
    dv_qt.add(f"E2:E{last_row}")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
